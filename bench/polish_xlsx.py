# Polish Excel ที่ parse_consensus/parse_image_consensus สร้าง ให้อ่านง่าย
#   wrap ทุกเซลล์ + header ตัวหนา/พื้น + freeze + FAIL แดง
#   --images : ฝัง "รูปจริง" ลงคอลัมน์ 'รูป' จาก path ในคอลัมน์ '_imgpath' (PNG/JPG/PDF)
# python bench/polish_xlsx.py [ไฟล์.xlsx] [--images]
import sys, os, openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

args = [a for a in sys.argv[1:] if not a.startswith('--')]
path = args[0] if args else 'bench/parse_consensus.xlsx'
do_images = '--images' in sys.argv

wb = openpyxl.load_workbook(path)
ws = wb['ByCase'] if 'ByCase' in wb.sheetnames else wb.worksheets[0]
headers = [str(c.value) for c in ws[1]]
col = lambda pred, dflt=None: next((i + 1 for i, h in enumerate(headers) if pred(h)), dflt)

# ── style พื้นฐาน ──
head = Font(bold=True); headfill = PatternFill('solid', fgColor='DDEBF7'); red = PatternFill('solid', fgColor='FCE4E4')
for c in ws[1]:
    c.font = head; c.fill = headfill
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical='top')
ws.freeze_panes = 'A2'
res_col = col(lambda h: h.strip() == 'ผล', 5)
for r in range(2, ws.max_row + 1):
    if ws.cell(r, res_col).value == 'FAIL':
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = red

# ── ฝังรูป ──
if do_images:
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    path_col = col(lambda h: '_imgpath' in h)
    img_col = col(lambda h: 'รูป' in h or 'image' in h.lower(), 2)
    thumbs = os.path.join(os.path.dirname(path) or '.', '.thumbs'); os.makedirs(thumbs, exist_ok=True)
    ws.column_dimensions[get_column_letter(img_col)].width = 40
    def load(p):                                   # PNG/JPG ตรงๆ, PDF เรนเดอร์หน้าแรก
        if p.lower().endswith('.pdf'):
            import fitz; d = fitz.open(p); pix = d[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            im = PILImage.frombytes('RGB', (pix.width, pix.height), pix.samples); d.close(); return im
        return PILImage.open(p).convert('RGB')
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, path_col).value if path_col else None
        if not p or not os.path.exists(p): continue
        try:
            im = load(p); im.thumbnail((260, 260))
            tp = os.path.join(thumbs, f'r{r}.png'); im.save(tp)
            xi = XLImage(tp); xi.anchor = f'{get_column_letter(img_col)}{r}'; ws.add_image(xi)
            ws.row_dimensions[r].height = im.height * 0.75 + 8
        except Exception as e:
            ws.cell(r, img_col).value = f'(รูปเปิดไม่ได้: {os.path.basename(p)})'
    if path_col:
        ws.column_dimensions[get_column_letter(path_col)].hidden = True

wb.save(path)
print(f'polished {path}: wrap/freeze/FAIL{" + ฝังรูป" if do_images else ""}')
